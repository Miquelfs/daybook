"""
EASA Part-FCL.050 logbook — Excel (.xlsx) export.

Sheet 1 "Logbook": all flight rows in EASA column order.
Sheet 2 "Stats": totals summary.
"""

import io
import sqlite3

from domains.aviation.aviation_config import EASA_COLUMNS


def _sec_to_hhmm(seconds: int | None) -> str:
    if not seconds:
        return ""
    h = seconds // 3600
    m = (seconds % 3600) // 60
    return f"{h}:{m:02d}"


def generate_excel(conn: sqlite3.Connection) -> bytes:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("openpyxl is required: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Logbook"

    # Header styling
    header_fill = PatternFill(start_color="0B3D5E", end_color="0B3D5E", fill_type="solid")
    header_font = Font(bold=True, color="FAFAFA", size=8)
    alt_fill = PatternFill(start_color="18181B", end_color="18181B", fill_type="solid")
    pic_fill = PatternFill(start_color="1A0D3D", end_color="1A0D3D", fill_type="solid")
    sim_fill = PatternFill(start_color="3D2200", end_color="3D2200", fill_type="solid")

    # Write header
    for col_idx, col_name in enumerate(EASA_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    flight_rows = conn.execute("""
        SELECT f.*,
               a1.name AS dep_name, a1.iata AS dep_iata_a, a1.country AS dep_country,
               a2.name AS arr_name, a2.iata AS arr_iata_a, a2.country AS arr_country
        FROM flights f
        LEFT JOIN airports a1 ON a1.icao = f.dep_icao
        LEFT JOIN airports a2 ON a2.icao = f.arr_icao
        WHERE f.is_sim = 0
        ORDER BY f.date, f.off_block_utc
    """).fetchall()

    sim_rows = conn.execute(
        "SELECT * FROM flights WHERE is_sim = 1 ORDER BY date"
    ).fetchall()

    row_num = 2
    for idx, r in enumerate(flight_rows):
        dep_time = r["off_block_utc"][11:16] if r["off_block_utc"] else ""
        arr_time = r["on_block_utc"][11:16] if r["on_block_utc"] else ""
        is_pic = r["crew_role"] == "pic"
        total = _sec_to_hhmm(r["block_seconds"])
        pic_t = _sec_to_hhmm(r["pic_seconds"])
        sic_t = _sec_to_hhmm(r["sic_seconds"])
        night_t = _sec_to_hhmm(r["night_seconds"])

        values = [
            r["date"],
            r["dep_icao"] or "",
            dep_time,
            r["arr_icao"] or "",
            arr_time,
            r["aircraft_type"] or "",
            r["aircraft_reg"] or "",
            "", "",                          # SP SE, SP ME
            total,                           # MP
            total,                           # Total
            "SELF" if is_pic else "",
            r["takeoffs_day"] or 0,
            r["takeoffs_night"] or 0,
            r["landings_day"] or 0,
            r["landings_night"] or 0,
            night_t,
            total,                           # IFR = block for MP ops
            pic_t,
            sic_t,
            "", "",                          # Dual, Instructor
            "", "", "",                      # FSTD Date/Type/Total
            r["notes"] or "",
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.font = Font(size=8)
            cell.alignment = Alignment(vertical="center")
            if is_pic:
                cell.fill = pic_fill
            elif idx % 2 == 1:
                cell.fill = alt_fill

        row_num += 1

    for r in sim_rows:
        values = [
            r["date"], "", "", "", "", "", "", "", "", "", "", "",
            0, 0, 0, 0, "", "", "", "", "", "",
            r["date"],
            r["sim_type"] or "",
            _sec_to_hhmm(r["block_seconds"]),
            r["notes"] or "",
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.font = Font(size=8)
            cell.fill = sim_fill
            cell.alignment = Alignment(vertical="center")
        row_num += 1

    # Auto-column widths (approximate)
    col_widths = [11, 7, 7, 7, 7, 14, 10, 5, 5, 6, 6, 8, 6, 6, 6, 6, 6, 6, 6, 7, 5, 6, 10, 10, 8, 20]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Sheet 2: Stats
    ws2 = wb.create_sheet("Stats")
    tot = conn.execute("""
        SELECT
            ROUND(SUM(block_seconds)/3600.0, 2) AS block_h,
            ROUND(SUM(pic_seconds)/3600.0, 2)   AS pic_h,
            ROUND(SUM(sic_seconds)/3600.0, 2)   AS sic_h,
            ROUND(SUM(night_seconds)/3600.0, 2) AS night_h,
            COUNT(*) AS sectors,
            SUM(takeoffs_day)+SUM(takeoffs_night) AS takeoffs,
            SUM(landings_day)+SUM(landings_night) AS landings,
            SUM(distance_nm) AS dist_nm
        FROM flights WHERE is_sim = 0
    """).fetchone()

    stats = [
        ("Total Block Hours", tot["block_h"] or 0),
        ("PIC Hours", tot["pic_h"] or 0),
        ("First Officer Hours", tot["sic_h"] or 0),
        ("Night Hours", tot["night_h"] or 0),
        ("Total Sectors", tot["sectors"] or 0),
        ("Total Takeoffs", tot["takeoffs"] or 0),
        ("Total Landings", tot["landings"] or 0),
        ("Distance (NM)", round(tot["dist_nm"] or 0, 0)),
    ]

    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 15

    hdr_font = Font(bold=True, color="FAFAFA", size=9)
    hdr_fill = PatternFill(start_color="0B3D5E", end_color="0B3D5E", fill_type="solid")

    for col, txt in [(1, "Metric"), (2, "Value")]:
        c = ws2.cell(row=1, column=col, value=txt)
        c.font = hdr_font
        c.fill = hdr_fill

    for row_i, (label, value) in enumerate(stats, 2):
        ws2.cell(row=row_i, column=1, value=label).font = Font(size=9)
        ws2.cell(row=row_i, column=2, value=value).font = Font(size=9)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
